import openpyxl
import re
import sys
import os
import datetime


#Чтение xlsx-файла
def read_file() -> dict:
	
	print('Имя файла или Enter(input.xlsx - по умолчанию)')
	file_name = input()
	if file_name == '':
		wb = openpyxl.load_workbook('inputt.xlsx')
	else:
		wb = openpyxl.load_workbook(file_name)
	
	
	print('Название листа или Enter(Лист1 - по умолчанию): ')
	sheet_name = input()
	
	if sheet_name == '':
		sheet = wb['Лист1']
	else:
		sheet = wb[str(sheet_name)]
	
	sheet_dict = {'sheet_obj': sheet,'max_row': sheet.max_row, 'max_column': sheet.max_column}
	
	return sheet_dict

#Преобразование xlsx-файла в словарь	
def compile_table() -> dict:
	sheet = read_file()
	header_alias = alias_dict()
	table_dict = {}
	
	for y in range(1,int(sheet.get('max_column')) + 1):
		for k, v in header_alias.items():
			regex = re.compile(v, re.IGNORECASE)
			reg_res = re.search(regex, sheet.get('sheet_obj').cell(row=1, column=y).value)
			if reg_res:
				table_header = str(reg_res.group(2)) + str(reg_res.group(3))
				table_header = table_header.upper()
				print(str(y) + ' Regex result: ' + str(reg_res.group(2)) + str(reg_res.group(3)))
				row_value_lst = []
				for i in range(2,int(sheet.get('max_row'))+1):
						temp = str(sheet.get('sheet_obj').cell(row=i, column=y).value)
						row_value_lst += [temp]
						table_dict[table_header] = row_value_lst
					
			#else:
				#print(str(y) + 'No match')

	table_dict['len_of_element_list'] = sheet.get('max_row') - 1
	return table_dict.copy()	

#Создание словаря шаблонов, по которым ищутся совпадения с параметрами команды
def alias_dict() -> dict:
	os.chdir('..')
	file = open('regex_template.txt', 'r')
	header_alias = {}

	for line in file:
		regex = re.compile(line, re.IGNORECASE)
		match_key = re.search("'(\w*)'",line)
		match_value = re.search('\(.*\)',line)

		try:
			rmatch_value = str(match_value.group())
			header_alias[match_key.group(1)] = rmatch_value.replace("\\\\",'\\')
		except AttributeError:
			pass
	
	return header_alias
	
#Создание скриптов на основе всех входных данных
def generate_script():
	
	table_dict = compile_table()
	input_string = get_string()
	element_length = table_dict['len_of_element_list']
	final_script_list = []
	index_of_final_list = []
	
	try:
		for k,v in table_dict.items():
			j = 0
			if str(k) in input_string:
				print(str(k))
				for i in range(element_length):
					if j in index_of_final_list:
						regex = re.escape(str(k)) + '=\-?\"?[A-Za-z0-9]*\"?'
						restring = re.sub(regex, re.escape(str(k)) + '=' + str(v[i]), final_script_list[j])
						print(restring)
						# input_string = str(restring)
						final_script_list[j] = restring

					else:
						regex = re.escape(str(k)) + '=\-?\"?[A-Za-z0-9]*\"?'
						restring = re.sub(regex, re.escape(str(k)) + '=' + str(v[i]), input_string)
						# input_string = str(restring)
						final_script_list.insert(j, restring)
						index_of_final_list += [j]
					j+=1

	except TypeError:
		pass
	count_file = 0
	os.chdir('./gen_scripts/')
	files_in_dir = os.listdir()
	while str('S_' + str(count_file) + '_' + str(get_output_name())) in files_in_dir:
		count_file += 1
	with open('S_' + str(count_file) + '_' + str(get_output_name()), 'w') as f:
		for i in range(len(final_script_list)):
			print(final_script_list[i])

			f.write(final_script_list[i] + '\n')
		print('OK')
	f.close

#Получить имя файла с готовыми командами
def get_output_name() -> str:#прописать исключение вместо if/else
	#c_date = 'default ' + str(datetime.date.today()) + '.txt'
	c_date = str(datetime.date.today()) + '.txt'
		
	return c_date

#Получить входную команду	
def get_string() -> str:
	print('\nКоманда: ')
	input_string = input()
	
	return input_string







#---------------main----------------
os.chdir('./files/')
generate_script()

